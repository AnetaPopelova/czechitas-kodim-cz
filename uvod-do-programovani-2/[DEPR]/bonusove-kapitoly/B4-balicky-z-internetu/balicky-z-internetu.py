## data do excelu

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
data = [
    {"nazev": "Projekt A", "hodiny": 550, "sazba-na-hodinu": 1000},
    {"nazev": "Projekt B", "hodiny": 470, "sazba-na-hodinu": 950},
    {"nazev": "Projekt C", "hodiny": 590, "sazba-na-hodinu": 1050},
]
row = 1
for radek in data:
    ws.cell(row=row, column=1).value = radek["nazev"]
    ws.cell(row=row, column=2).value = radek["hodiny"]
    ws.cell(row=row, column=3).value = radek["sazba-na-hodinu"]
    row += 1
wb.save("soubor.xlsx")
